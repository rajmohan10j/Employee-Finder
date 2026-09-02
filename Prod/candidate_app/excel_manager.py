import os
import shutil
import threading
from datetime import datetime
from pathlib import Path
import openpyxl

from backup_manager import BackupManager

TRACKER_HEADERS = [
    'Candidate Name',
    'Phone Number',
    'Email',
    'Location',
    'Total Experience',
    'Current Position / Role',
    'Domain / Industry',
    'Education Background',
    'Open To Work / Active',
    'Portal Source',
    'PDF File Name',
    'Processed Timestamp',
    'Resume File Name',
    'HR Called',
    'Date',
    'HR Remarks',
    'Follow-up Date',
    'HR Follow-up Remarks',
    'Escalation Level / Person',
    'Escalation Action Category',
    'Escalation Remarks',
    # Conversion Intelligence Tracking Columns (added for analytics)
    'Call Response',
    'Interview / Meeting Agreed',
    'Advisory Role Interest',
    # Audience Segmentation (P1/P2 Insurance Advisory)
    'Age',
    'Employment Sector',
    'Retirement Status'
]


# Canonical status classification logic shared across /api/stats and candidate filtering
def is_candidate_closed(c: dict) -> bool:
    hr = (c.get("HR Called") or "").lower()
    otw = (c.get("Open To Work / Active") or "").lower()
    return ("not interested" in hr or "closed" in hr or "not interested" in otw or "closed" in otw)

def is_candidate_called(c: dict) -> bool:
    hr = (c.get("HR Called") or "").lower()
    return "yes" in hr and not is_candidate_closed(c)

def is_candidate_busy(c: dict) -> bool:
    hr = (c.get("HR Called") or "").lower()
    return ("busy" in hr or "call later" in hr or "call back" in hr) and not is_candidate_closed(c)

def is_candidate_not_reachable(c: dict) -> bool:
    hr = (c.get("HR Called") or "").lower()
    return ("not reachable" in hr or "rnr" in hr or "not connected" in hr) and not is_candidate_closed(c)

def is_candidate_pending(c: dict) -> bool:
    return (
        not is_candidate_called(c)
        and not is_candidate_closed(c)
        and not is_candidate_busy(c)
        and not is_candidate_not_reachable(c)
    )

def is_candidate_followup(c: dict) -> bool:
    if is_candidate_closed(c):
        return False
    f_date = (c.get("Follow-up Date") or "").strip()
    f_rem = (c.get("HR Follow-up Remarks") or "").strip().lower()
    if f_date and f_date.lower() not in ["", "none", "n/a", "null"]:
        return True
    if f_rem and f_rem.lower() not in ["", "none", "n/a", "null", "not interested", "closed"]:
        return True
    return False

def is_candidate_assigned(c: dict) -> bool:
    esc = (c.get("Escalation Level / Person") or "").strip()
    return bool(esc and esc.lower() not in ["", "none", "no escalation", "none / no escalation", "unassigned"])


class ExcelManager:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path).resolve()
        self.lock = threading.RLock() # Re-entrant lock prevents deadlocks on nested calls
        self.backup_mgr = BackupManager(str(self.file_path))
        
        # Take pre-connection / startup backup before reading or modifying workbook
        if self.file_path.exists():
            self.backup_mgr.create_backup(tier='sessions', prefix='startup')

        self._ensure_file_exists()

    def _ensure_file_exists(self):
        if not self.file_path.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tracker"
            ws.append(TRACKER_HEADERS)
            wb.save(self.file_path)
        else:
            # Upgrade headers in existing file if new columns added
            with self.lock:
                try:
                    wb = openpyxl.load_workbook(self.file_path)
                    ws = wb.active
                    existing_headers = [str(ws.cell(row=1, column=c).value).strip() for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value]
                    changed = False
                    for idx, header in enumerate(TRACKER_HEADERS):
                        if idx >= len(existing_headers) or existing_headers[idx] != header:
                            ws.cell(row=1, column=idx + 1).value = header
                            changed = True
                    if changed:
                        wb.save(self.file_path)
                except Exception as e:
                    print(f"[Header Sync Warning] {e}")

    def _create_backup(self, prefix: str = 'edit'):
        """Route to enterprise BackupManager"""
        try:
            return self.backup_mgr.create_backup(tier='sessions', prefix=prefix)
        except Exception as e:
            print(f"[Backup Error] {e}")
            return None

    def get_all_candidates(self, query=None, filter_status=None, filter_portal=None, filter_escalation=None):
        with self.lock:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws = wb.active
            
            # Identify headers from Row 1
            headers = []
            for col in range(1, max(len(TRACKER_HEADERS), ws.max_column) + 1):
                val = ws.cell(row=1, column=col).value
                if val:
                    headers.append(str(val).strip())
                elif col <= len(TRACKER_HEADERS):
                    headers.append(TRACKER_HEADERS[col-1])

            candidates = []
            for row_idx in range(2, ws.max_row + 1):
                row_data = {"_row_id": row_idx}
                has_any_data = False
                for col_idx, header in enumerate(headers):
                    cell_val = ws.cell(row=row_idx, column=col_idx + 1).value
                    if cell_val is not None and str(cell_val).strip() != "":
                        has_any_data = True
                        row_data[header] = str(cell_val).strip()
                    else:
                        row_data[header] = ""
                
                if not has_any_data:
                    continue

                # Filter logic
                if query:
                    q = query.lower()
                    matched = any(q in str(v).lower() for k, v in row_data.items() if k != "_row_id")
                    if not matched:
                        continue

                if filter_status and filter_status != "All":
                    fs = filter_status.lower()
                    if fs == "called":
                        if not is_candidate_called(row_data):
                            continue
                    elif fs == "pending":
                        if not is_candidate_pending(row_data):
                            continue
                    elif fs in ["closed", "not_interested", "not interested"]:
                        if not is_candidate_closed(row_data):
                            continue
                    elif fs in ["busy", "call_later"]:
                        if not is_candidate_busy(row_data):
                            continue
                    elif fs in ["not_reachable", "rnr"]:
                        if not is_candidate_not_reachable(row_data):
                            continue
                    elif fs in ["followups", "follow_ups", "follow-ups"]:
                        if not is_candidate_followup(row_data):
                            continue

                if filter_portal and filter_portal != "All":
                    portal = row_data.get("Portal Source", "").lower()
                    if filter_portal.lower() not in portal:
                        continue

                if filter_escalation and filter_escalation != "All":
                    esc = (row_data.get("Escalation Level / Person") or "").strip()
                    esc_lower = esc.lower()
                    fe_lower = filter_escalation.lower().strip()
                    if fe_lower in ["none", "no escalation", "unassigned"]:
                        if is_candidate_assigned(row_data):
                            continue
                    elif fe_lower in ["assigned", "escalated", "has_task", "all tasks", "action required"]:
                        if not is_candidate_assigned(row_data):
                            continue
                    elif fe_lower not in esc_lower:
                        continue

                candidates.append(row_data)

            return candidates

    def get_candidate_by_id(self, row_id: int):
        with self.lock:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws = wb.active
            
            headers = [TRACKER_HEADERS[c-1] for c in range(1, len(TRACKER_HEADERS) + 1)]
            for col in range(1, len(TRACKER_HEADERS) + 1):
                val = ws.cell(row=1, column=col).value
                if val:
                    headers[col-1] = str(val).strip()

            if row_id < 2 or row_id > ws.max_row:
                return None

            candidate = {"_row_id": row_id}
            for col_idx, header in enumerate(headers):
                cell_val = ws.cell(row=row_id, column=col_idx + 1).value
                candidate[header] = str(cell_val).strip() if cell_val is not None else ""

            return candidate

    def update_candidate(self, row_id: int, data: dict):
        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active

            headers = [TRACKER_HEADERS[c-1] for c in range(1, len(TRACKER_HEADERS) + 1)]
            for col in range(1, len(TRACKER_HEADERS) + 1):
                val = ws.cell(row=1, column=col).value
                if val:
                    headers[col-1] = str(val).strip()

            if row_id < 2 or row_id > ws.max_row:
                raise ValueError(f"Candidate row ID {row_id} does not exist.")

            for col_idx, header in enumerate(headers):
                if header in data:
                    val = data[header]
                    ws.cell(row=row_id, column=col_idx + 1).value = val if val is not None else ""

            wb.save(self.file_path)
            return self.get_candidate_by_id(row_id)

    def add_candidate(self, data: dict):
        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active

            headers = [TRACKER_HEADERS[c-1] for c in range(1, len(TRACKER_HEADERS) + 1)]
            for col in range(1, len(TRACKER_HEADERS) + 1):
                val = ws.cell(row=1, column=col).value
                if val:
                    headers[col-1] = str(val).strip()

            new_row_idx = ws.max_row + 1
            if not data.get("Processed Timestamp"):
                data["Processed Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            for col_idx, header in enumerate(headers):
                val = data.get(header, "")
                ws.cell(row=new_row_idx, column=col_idx + 1).value = val if val is not None else ""

            wb.save(self.file_path)
            return self.get_candidate_by_id(new_row_idx)

    def delete_candidate(self, row_id: int):
        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active

            if row_id < 2 or row_id > ws.max_row:
                raise ValueError(f"Candidate row ID {row_id} does not exist.")

            ws.delete_rows(row_id, 1)
            wb.save(self.file_path)
            return True

    def import_excel_file(self, import_source, mode='append'):
        """
        Import candidates from an external .xlsx file.
        import_source can be a filepath (str/Path) or file-like object (stream).
        mode: 'append' to add to existing, 'replace' to overwrite all records.
        """
        with self.lock:
            self._create_backup()
            
            # Load incoming workbook
            in_wb = openpyxl.load_workbook(import_source, data_only=True)
            in_ws = in_wb.active

            # Find matching headers in imported sheet
            in_headers = {}
            for col in range(1, in_ws.max_column + 1):
                val = in_ws.cell(row=1, column=col).value
                if val:
                    clean_val = str(val).strip().lower().replace('_', ' ').replace('-', ' ')
                    in_headers[col] = clean_val

            # Header mapping dictionary
            header_map = {
                'name': 'Candidate Name',
                'candidate name': 'Candidate Name',
                'candidate': 'Candidate Name',
                'title': 'Candidate Name',
                'phone': 'Phone Number',
                'phone number': 'Phone Number',
                'contact': 'Phone Number',
                'mobile': 'Phone Number',
                'email': 'Email',
                'email id': 'Email',
                'email address': 'Email',
                'location': 'Location',
                'city': 'Location',
                'experience': 'Total Experience',
                'total experience': 'Total Experience',
                'exp': 'Total Experience',
                'open to work': 'Open To Work / Active',
                'open to work / active': 'Open To Work / Active',
                'status': 'Open To Work / Active',
                'portal': 'Portal Source',
                'portal source': 'Portal Source',
                'source': 'Portal Source',
                'pdf file name': 'PDF File Name',
                'pdf': 'PDF File Name',
                'processed timestamp': 'Processed Timestamp',
                'timestamp': 'Processed Timestamp',
                'date': 'Date',
                'resume file name': 'Resume File Name',
                'resume': 'Resume File Name',
                'hr called': 'HR Called',
                'called': 'HR Called',
                'hr response': 'HR Remarks',
                'response': 'HR Remarks',
                'hr remarks': 'HR Remarks',
                'remarks': 'HR Remarks',
                'notes': 'HR Remarks',
                'feedback': 'HR Remarks',
                'follow up date': 'Follow-up Date',
                'followup date': 'Follow-up Date',
                'follow-up date': 'Follow-up Date',
                'hr follow up remarks': 'HR Follow-up Remarks',
                'follow up remarks': 'HR Follow-up Remarks',
                'followup remarks': 'HR Follow-up Remarks',
                'escalation level / person': 'Escalation Level / Person',
                'escalation level person': 'Escalation Level / Person',
                'escalation level': 'Escalation Level / Person',
                'escalation person': 'Escalation Level / Person',
                'escalated to': 'Escalation Level / Person',
                'task / assigned to': 'Escalation Level / Person',
                'task assigned to': 'Escalation Level / Person',
                'assigned to': 'Escalation Level / Person',
                'escalation action category': 'Escalation Action Category',
                'escalation action': 'Escalation Action Category',
                'action category': 'Escalation Action Category',
                'action': 'Escalation Action Category',
                'escalation remarks': 'Escalation Remarks',
                'escalation comments': 'Escalation Remarks',
                'escalation notes': 'Escalation Remarks',
                'escalation details': 'Escalation Remarks'
            }

            # Open target master workbook
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active

            if mode == 'replace':
                # Delete all rows except header
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row - 1)

            # Master column indices
            master_cols = {h: idx + 1 for idx, h in enumerate(TRACKER_HEADERS)}

            imported_count = 0
            curr_row = ws.max_row + 1

            for r in range(2, in_ws.max_row + 1):
                row_has_data = False
                row_dict = {h: "" for h in TRACKER_HEADERS}
                
                for col_idx, col_name in in_headers.items():
                    cell_val = in_ws.cell(row=r, column=col_idx).value
                    if cell_val is not None and str(cell_val).strip() != "":
                        row_has_data = True
                        mapped_target = header_map.get(col_name)
                        if mapped_target:
                            row_dict[mapped_target] = str(cell_val).strip()

                if not row_has_data:
                    continue

                if not row_dict['Processed Timestamp']:
                    row_dict['Processed Timestamp'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                for header_name, target_col in master_cols.items():
                    ws.cell(row=curr_row, column=target_col).value = row_dict.get(header_name, "")

                curr_row += 1
                imported_count += 1

            wb.save(self.file_path)
            return {
                "imported_count": imported_count,
                "total_count": ws.max_row - 1
            }
