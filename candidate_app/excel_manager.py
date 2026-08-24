import os
import shutil
import threading
from datetime import datetime
from pathlib import Path
import openpyxl

TRACKER_HEADERS = [
    'Candidate Name',
    'Phone Number',
    'Email',
    'Location',
    'Total Experience',
    'Open To Work / Active',
    'Portal Source',
    'PDF File Name',
    'Processed Timestamp',
    'Resume File Name',
    'HR Called',
    'Date',
    'HR Remarks',
    'Follow-up Date',
    'HR Follow-up Remarks'
]

class ExcelManager:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path).resolve()
        self.lock = threading.RLock() # Re-entrant lock prevents deadlocks on nested calls
        self.backup_dir = self.file_path.parent / "backups"
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        self._ensure_file_exists()

    def _ensure_file_exists(self):
        if not self.file_path.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tracker"
            ws.append(TRACKER_HEADERS)
            wb.save(self.file_path)

    def _create_backup(self):
        try:
            if self.file_path.exists():
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_file = self.backup_dir / f"candidates_tracker_{timestamp}.xlsx"
                shutil.copy2(self.file_path, backup_file)
                # Keep only last 20 backups
                backups = sorted(self.backup_dir.glob("candidates_tracker_*.xlsx"), key=os.path.getmtime)
                if len(backups) > 20:
                    for old_b in backups[:-20]:
                        try:
                            old_b.unlink()
                        except Exception:
                            pass
        except Exception as e:
            print(f"[Backup Error] {e}")

    def get_all_candidates(self, query=None, filter_status=None, filter_portal=None):
        with self.lock:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws = wb.active
            
            # Identify headers from Row 1
            headers = []
            for col in range(1, len(TRACKER_HEADERS) + 1):
                val = ws.cell(row=1, column=col).value
                headers.append(str(val).strip() if val else TRACKER_HEADERS[col-1])

            candidates = []
            for row_idx in range(2, ws.max_row + 1):
                row_data = {"_row_id": row_idx}
                has_any_data = False
                for col_idx, header in enumerate(headers):
                    cell_val = ws.cell(row=row_idx, column=col_idx + 1).value
                    if cell_val is not None and str(cell_val).strip() != "":
                        has_any_data = True
                        row_data[header] = str(cell_val).strip()
                    else:
                        row_data[header] = ""
                
                if not has_any_data:
                    continue

                # Filter logic
                if query:
                    q = query.lower()
                    matched = any(q in str(v).lower() for k, v in row_data.items() if k != "_row_id")
                    if not matched:
                        continue

                if filter_status and filter_status != "All":
                    hr_called = row_data.get("HR Called", "").lower()
                    if filter_status.lower() == "called" and "yes" not in hr_called:
                        continue
                    if filter_status.lower() == "pending" and ("yes" in hr_called):
                        continue

                if filter_portal and filter_portal != "All":
                    portal = row_data.get("Portal Source", "").lower()
                    if filter_portal.lower() not in portal:
                        continue

                candidates.append(row_data)

            return candidates

    def get_candidate_by_id(self, row_id: int):
        with self.lock:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws = wb.active
            
            headers = [TRACKER_HEADERS[c-1] for c in range(1, len(TRACKER_HEADERS) + 1)]
            for col in range(1, len(TRACKER_HEADERS) + 1):
                val = ws.cell(row=1, column=col).value
                if val:
                    headers[col-1] = str(val).strip()

            if row_id < 2 or row_id > ws.max_row:
                return None

            candidate = {"_row_id": row_id}
            for col_idx, header in enumerate(headers):
                cell_val = ws.cell(row=row_id, column=col_idx + 1).value
                candidate[header] = str(cell_val).strip() if cell_val is not None else ""

            return candidate

    def update_candidate(self, row_id: int, data: dict):
        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active

            headers = [TRACKER_HEADERS[c-1] for c in range(1, len(TRACKER_HEADERS) + 1)]
            for col in range(1, len(TRACKER_HEADERS) + 1):
                val = ws.cell(row=1, column=col).value
                if val:
                    headers[col-1] = str(val).strip()

            if row_id < 2 or row_id > ws.max_row:
                raise ValueError(f"Candidate row ID {row_id} does not exist.")

            for col_idx, header in enumerate(headers):
                if header in data:
                    val = data[header]
                    ws.cell(row=row_id, column=col_idx + 1).value = val if val is not None else ""

            wb.save(self.file_path)
            return self.get_candidate_by_id(row_id)

    def add_candidate(self, data: dict):
        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active

            headers = [TRACKER_HEADERS[c-1] for c in range(1, len(TRACKER_HEADERS) + 1)]
            for col in range(1, len(TRACKER_HEADERS) + 1):
                val = ws.cell(row=1, column=col).value
                if val:
                    headers[col-1] = str(val).strip()

            new_row_idx = ws.max_row + 1
            if not data.get("Processed Timestamp"):
                data["Processed Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            for col_idx, header in enumerate(headers):
                val = data.get(header, "")
                ws.cell(row=new_row_idx, column=col_idx + 1).value = val if val is not None else ""

            wb.save(self.file_path)
            return self.get_candidate_by_id(new_row_idx)

    def delete_candidate(self, row_id: int):
        with self.lock:
            self._create_backup()
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active

            if row_id < 2 or row_id > ws.max_row:
                raise ValueError(f"Candidate row ID {row_id} does not exist.")

            ws.delete_rows(row_id, 1)
            wb.save(self.file_path)
            return True

    def import_excel_file(self, import_source, mode='append'):
        """
        Import candidates from an external .xlsx file.
        import_source can be a filepath (str/Path) or file-like object (stream).
        mode: 'append' to add to existing, 'replace' to overwrite all records.
        """
        with self.lock:
            self._create_backup()
            
            # Load incoming workbook
            in_wb = openpyxl.load_workbook(import_source, data_only=True)
            in_ws = in_wb.active

            # Find matching headers in imported sheet
            in_headers = {}
            for col in range(1, in_ws.max_column + 1):
                val = in_ws.cell(row=1, column=col).value
                if val:
                    clean_val = str(val).strip().lower().replace('_', ' ').replace('-', ' ')
                    in_headers[col] = clean_val

            # Header mapping dictionary
            header_map = {
                'name': 'Candidate Name',
                'candidate name': 'Candidate Name',
                'candidate': 'Candidate Name',
                'title': 'Candidate Name',
                'phone': 'Phone Number',
                'phone number': 'Phone Number',
                'contact': 'Phone Number',
                'mobile': 'Phone Number',
                'email': 'Email',
                'email id': 'Email',
                'email address': 'Email',
                'location': 'Location',
                'city': 'Location',
                'experience': 'Total Experience',
                'total experience': 'Total Experience',
                'exp': 'Total Experience',
                'open to work': 'Open To Work / Active',
                'open to work / active': 'Open To Work / Active',
                'status': 'Open To Work / Active',
                'portal': 'Portal Source',
                'portal source': 'Portal Source',
                'source': 'Portal Source',
                'pdf file name': 'PDF File Name',
                'pdf': 'PDF File Name',
                'processed timestamp': 'Processed Timestamp',
                'timestamp': 'Processed Timestamp',
                'date': 'Date',
                'resume file name': 'Resume File Name',
                'resume': 'Resume File Name',
                'hr called': 'HR Called',
                'called': 'HR Called',
                'hr response': 'HR Remarks',
                'response': 'HR Remarks',
                'hr remarks': 'HR Remarks',
                'remarks': 'HR Remarks',
                'notes': 'HR Remarks',
                'feedback': 'HR Remarks',
                'follow up date': 'Follow-up Date',
                'followup date': 'Follow-up Date',
                'follow-up date': 'Follow-up Date',
                'hr follow up remarks': 'HR Follow-up Remarks',
                'follow up remarks': 'HR Follow-up Remarks',
                'followup remarks': 'HR Follow-up Remarks'
            }

            # Open target master workbook
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active

            if mode == 'replace':
                # Delete all rows except header
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row - 1)

            # Master column indices
            master_cols = {h: idx + 1 for idx, h in enumerate(TRACKER_HEADERS)}

            imported_count = 0
            curr_row = ws.max_row + 1

            for r in range(2, in_ws.max_row + 1):
                row_has_data = False
                row_dict = {h: "" for h in TRACKER_HEADERS}
                
                for col_idx, col_name in in_headers.items():
                    cell_val = in_ws.cell(row=r, column=col_idx).value
                    if cell_val is not None and str(cell_val).strip() != "":
                        row_has_data = True
                        mapped_target = header_map.get(col_name)
                        if mapped_target:
                            row_dict[mapped_target] = str(cell_val).strip()

                if not row_has_data:
                    continue

                if not row_dict['Processed Timestamp']:
                    row_dict['Processed Timestamp'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                for header_name, target_col in master_cols.items():
                    ws.cell(row=curr_row, column=target_col).value = row_dict.get(header_name, "")

                curr_row += 1
                imported_count += 1

            wb.save(self.file_path)
            return {
                "imported_count": imported_count,
                "total_count": ws.max_row - 1
            }
